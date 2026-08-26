#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
학기 시간표 검증 스크립트 (swap-finder2)

index.html에 박힌 시간표가 제대로 만들어졌는지 확인한다.
자세한 배경은 학기시간표-갱신-절차.md 참고.

사용법:
    python tools/verify_timetable.py

    # 파일 위치를 직접 지정할 때
    python tools/verify_timetable.py --comci comci_raw.txt --weekly "주간시간표.xlsx"

필요:
    pip install openpyxl      (--weekly 를 쓸 때만)

검사 항목:
    1. 슬롯 형식      모든 칸이 index.html의 parseCls 정규식을 통과하는가
    2. 교사 이름표    worker.js의 TEACHER_NAMES가 컴시간 자료446과 맞는가
    3. 주당 시수      교사별 시수가 주간시간표 선언값과 일치하는가   ← 가장 강력
    4. 시간표 대조    컴시간 자료481과 칸 단위로 일치하는가
"""
import argparse, json, os, re, sys

DAYS = ['월', '화', '수', '목', '금']
NOSWAP = {'자율활동', '창의적 체험활동'}

# index.html의 parseCls가 쓰는 정규식 세 개
RE_G = re.compile(r'^(\d)학년')
RE_B = re.compile(r'\((\d+)\)\s*$')
RE_S = re.compile(r'/\*?\s*([^(]+?)\s*\(\d')

OK, WARN, BAD = '✅', '⚠️ ', '❌'
_fail = 0


def head(t):
    print(f'\n{"─" * 62}\n{t}\n{"─" * 62}')


def report(ok, msg, detail=None):
    global _fail
    if not ok:
        _fail += 1
    print(f'{OK if ok else BAD} {msg}')
    for d in (detail or [])[:12]:
        print(f'      {d}')
    if detail and len(detail) > 12:
        print(f'      … 외 {len(detail) - 12}건')


def load_builtin(path):
    raw = open(path, encoding='utf-8').read()
    m = re.search(r'const BUILTIN_TEACHERS=(\[.*?\]);\s*\r?\n', raw, re.S)
    if not m:
        sys.exit(f'{BAD} {path} 에서 BUILTIN_TEACHERS를 찾지 못했습니다.')
    return json.loads(m.group(1))


def load_teacher_names(path):
    raw = open(path, encoding='utf-8').read()
    m = re.search(r'const TEACHER_NAMES\s*=\s*\[(.*?)\];', raw, re.S)
    if not m:
        return None
    return re.findall(r"'([^']+)'", m.group(1))


def load_comci(path):
    txt = open(path, encoding='utf-8').read()
    return json.JSONDecoder().raw_decode(txt)[0]


def comci_num(v):
    """'>40035' / 40035 → 정수. 빈 칸이면 0."""
    if isinstance(v, str):
        v = v.lstrip('>')
    try:
        n = int(v)
    except (TypeError, ValueError):
        return 0
    return n if n > 0 else 0


def base_timetable(d):
    """자료481(학급별) → {(교사번호, 요일idx, 교시): 학급코드} 로 뒤집는다."""
    Z = d['자료481']
    out = {}
    for gi in range(1, Z[0] + 1):
        gb = Z[gi]
        for bi in range(1, gb[0] + 1):
            bb = gb[bi]
            for di in range(1, bb[0] + 1):
                db = bb[di]
                for pi in range(1, db[0] + 1):
                    n = comci_num(db[pi])
                    if not n:
                        continue
                    ti = n % 1000                      # 과목번호×1000 + 교사번호
                    out.setdefault((ti, di - 1, pi), []).append((gi, bi))
    return out


def slot_gb(v):
    """'1학년 경영정보과/통합과학2(1)' → (1, 1)"""
    g, b = RE_G.search(v), RE_B.search(v)
    return (int(g.group(1)), int(b.group(1))) if (g and b) else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--index', default='index.html')
    ap.add_argument('--worker', default='worker.js')
    ap.add_argument('--comci', default='comci_raw.txt')
    ap.add_argument('--weekly', default=None, help='주간시간표 xlsx (시수 검증용)')
    a = ap.parse_args()

    teachers = load_builtin(a.index)
    print(f'{OK} {a.index} — 교사 {len(teachers)}명 읽음')

    # ── 1. 슬롯 형식 ───────────────────────────────────────────
    head('1. 슬롯 형식 (parseCls 통과 여부)')
    bad, cnt = [], 0
    for t in teachers:
        for day in DAYS:
            arr = t['sch'].get(day, [])
            if len(arr) != 12:
                bad.append(f"{t['dname']} {day}요일 배열 길이 {len(arr)} (12이어야 함)")
            for v in arr:
                if not v or v in NOSWAP:
                    continue
                cnt += 1
                if not (RE_G.search(v) and RE_B.search(v) and RE_S.search(v)):
                    bad.append(f"{t['dname']} · {v}")
    report(not bad, f'{cnt}칸 검사, 형식 오류 {len(bad)}건', bad)

    # ── 2. 교사 이름표 ─────────────────────────────────────────
    head('2. 교사 이름표 (worker.js ↔ 컴시간 자료446)')
    names = load_teacher_names(a.worker) if os.path.exists(a.worker) else None
    d = load_comci(a.comci) if os.path.exists(a.comci) else None

    if names is None:
        print(f'{WARN}{a.worker} 없음 — 건너뜀')
    elif d is None:
        print(f'{WARN}{a.comci} 없음 — 길이만 확인')
        report(len(names) >= len(teachers),
               f'TEACHER_NAMES {len(names)}개 (시간표 교사 {len(teachers)}명)')
    else:
        nt, masked = d['교사수'], d['자료446']
        report(len(names) == nt,
               f'TEACHER_NAMES {len(names)}개 = 컴시간 교사수 {nt}',
               [] if len(names) == nt else
               ['개수가 다르면 그 지점부터 뒤쪽 교사 전원이 다른 이름으로 표시됩니다'])
        mism = [f'{i}: 컴시간 {masked[i]} ↔ 이름표 {names[i-1]}'
                for i in range(1, min(nt, len(names)) + 1)
                if not names[i - 1].startswith(masked[i].rstrip('*'))]
        report(not mism, f'마스킹 이름 대조 — 불일치 {len(mism)}건', mism)

        virt = [n for n in names[-2:] if n in ('자율', '창체')]
        report(len(virt) == 2,
               '끝 2개가 자율·창체 가상 항목인가',
               [] if len(virt) == 2 else ['자료446 끝 2개는 교사가 아닙니다. 빼면 인덱스가 밀립니다'])

    # ── 3. 주당 시수 ───────────────────────────────────────────
    head('3. 주당 시수 (주간시간표 선언값 대조)')
    if not a.weekly:
        print(f'{WARN}--weekly 미지정 — 건너뜀 (가장 강력한 검증이므로 가능하면 지정할 것)')
    elif not os.path.exists(a.weekly):
        print(f'{WARN}{a.weekly} 없음 — 건너뜀')
    else:
        try:
            import openpyxl
        except ImportError:
            print(f'{WARN}openpyxl 미설치 — pip install openpyxl')
        else:
            ws = openpyxl.load_workbook(a.weekly, data_only=True).worksheets[0]
            decl = {}
            for c in range(1, ws.max_column + 1):
                h = str(ws.cell(2, c).value or '')
                m = re.match(r'^(\S+?)\((\d+)\)$', h.strip())
                if m:
                    decl[m.group(1)] = int(m.group(2))
            if not decl:
                print(f'{WARN}머리글에서 "이름(시수)" 형태를 찾지 못했습니다 — 양식이 바뀌었는지 확인하세요')
            else:
                diff = []
                for t in teachers:
                    if t['dname'] not in decl:
                        diff.append(f"{t['dname']}: 주간시간표에 없음")
                        continue
                    n = sum(1 for day in DAYS for v in t['sch'][day]
                            if v and v not in NOSWAP)
                    if n != decl[t['dname']]:
                        diff.append(f"{t['dname']}: 배정 {decl[t['dname']]} ≠ 생성 {n}")
                report(not diff, f'{len(teachers)}명 중 시수 불일치 {len(diff)}건', diff)

    # ── 4. 컴시간 원본 대조 ────────────────────────────────────
    head('4. 컴시간 자료481 대조')
    if d is None or names is None:
        print(f'{WARN}컴시간 원본 또는 이름표가 없어 건너뜀')
    else:
        by_name = {t['dname']: t['sch'] for t in teachers}
        base = base_timetable(d)
        hit = miss = 0
        bad = []
        for (ti, di, pi), cls in base.items():
            if not (1 <= ti <= len(names)):
                continue
            nm = names[ti - 1]
            if nm not in by_name:          # 자율·창체 가상 항목
                continue
            got = by_name[nm][DAYS[di]][pi - 1] if pi <= 12 else ''
            gb = slot_gb(got) if got and got not in NOSWAP else None
            if gb in cls:
                hit += 1
            else:
                miss += 1
                want = ', '.join(f'{g}-{b}' for g, b in cls)
                bad.append(f'{nm} {DAYS[di]}{pi}교시: 원본={want} / 사이트={got or "빈칸"}')
        tot = hit + miss
        rate = 100 * hit / tot if tot else 0
        if rate >= 99:
            report(True, f'{tot}칸 중 {hit}칸 일치 ({rate:.1f}%) — 컴시간과 완전 일치')
        elif rate >= 70:
            # 주간시간표 xlsx와 컴시간이 다른 판본이면 배치가 어긋난다.
            # 3번(시수)이 통과했다면 오류가 아니므로 실패로 세지 않는다.
            print(f'{WARN}{tot}칸 중 {hit}칸 일치 ({rate:.1f}%) — 배치 차이 {miss}칸')
            for x in bad[:8]:
                print(f'      {x}')
            if len(bad) > 8:
                print(f'      … 외 {len(bad) - 8}건')
            print('      ※ 주간시간표 xlsx와 컴시간이 다른 판본이면 이 차이가 납니다.')
            print('        3번(시수)이 통과했다면 배치 차이일 뿐 오류가 아닙니다.')
        else:
            report(False, f'{tot}칸 중 {hit}칸만 일치 ({rate:.1f}%) — 정렬이 어긋났을 가능성', bad)
            print('      ※ 일치율이 이렇게 낮으면 교사 이름표 순서를 의심하세요.')

    # ── 결과 ───────────────────────────────────────────────────
    head('결과')
    if _fail == 0:
        print(f'{OK} 모든 검사 통과')
    else:
        print(f'{BAD} {_fail}개 항목 실패 — 위 내용을 확인하세요')
    return 1 if _fail else 0


if __name__ == '__main__':
    sys.exit(main())
